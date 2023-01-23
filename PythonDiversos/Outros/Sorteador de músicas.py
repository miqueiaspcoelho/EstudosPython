import random as rd
import openpyxl
import time
import datetime

def pegaColuna(nome_file,i,j):
    arquivo = openpyxl.load_workbook('.\\'+nome_file) #carrega o arquivo
    folha=arquivo.active
    maximo_linhas=folha.max_row
    lista=[]
    for linha in range(i,maximo_linhas+1):
        celula=folha.cell(row=linha,column=j)
        lista.append(celula.value)
    return lista

def musicasElegiveis(data,musica,dias):
    lista=[]
    compara_data=0
    for u in range(0,len(data)):
        if data[u]!=None:
            compara_data=datetime.datetime.now()- data[u]
            if compara_data > datetime.timedelta(days=dias):
                lista.append(musica[u])
        else:
            lista.append(musica[u])
    return lista

dias=15
musica=pegaColuna('Repertório Exaltai.xlsx',2,5)
data=pegaColuna('Repertório Exaltai.xlsx',2,8)
musicas_elegiveis=musicasElegiveis(data,musica,dias)

y=''
sorteados=[]
c=int(input('Quantidade de números: '))

print('Músicas elegiveis:\n',musicas_elegiveis, len(musicas_elegiveis))
for x in range (0,c):
    y=rd.choice(musicas_elegiveis)
    musicas_elegiveis.remove(y)
    sorteados.append(y)
print('Músicas sorteadas:\n',sorteados)




